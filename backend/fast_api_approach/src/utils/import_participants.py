"""
Participant import helpers for CSV and Excel files.

This module provides simple parsing functions for importing participants
from CSV or Excel (XLSX) files into the database.
"""
import csv
import io
from typing import List, Dict, Any, Optional


def normalize_header(header: str) -> str:
    """
    Normalize a header string for matching.
    Strips whitespace and converts to lowercase.
    """
    if header is None:
        return ""
    return str(header).strip().lower()


def parse_csv_participants(file_content: bytes) -> List[Dict[str, Any]]:
    """
    Parse a CSV file and return a list of participant dictionaries.
    
    Expected columns (case-insensitive):
        - name
        - email
        - phone (optional)
    
    Args:
        file_content: Raw bytes of the CSV file
        
    Returns:
        List of dicts, each with keys: 'name', 'email', 'phone' (if present)
    """
    # Decode bytes to string (try utf-8, fallback to latin-1)
    try:
        text = file_content.decode('utf-8')
    except UnicodeDecodeError:
        text = file_content.decode('latin-1')
    
    # Use csv.DictReader to parse
    reader = csv.DictReader(io.StringIO(text))
    
    participants = []
    for row in reader:
        # Normalize headers for matching
        normalized_row = {normalize_header(k): v for k, v in row.items()}
        
        participant = {
            'name': normalized_row.get('name', '').strip() if normalized_row.get('name') else None,
            'email': normalized_row.get('email', '').strip() if normalized_row.get('email') else None,
            'phone': normalized_row.get('phone', '').strip() if normalized_row.get('phone') else None,
        }
        participants.append(participant)
    
    return participants


def parse_excel_participants(file_content: bytes) -> List[Dict[str, Any]]:
    """
    Parse an Excel (XLSX) file and return a list of participant dictionaries.
    Uses the first sheet, with the first row as headers.
    
    Expected columns (case-insensitive):
        - name
        - email
        - phone (optional)
    
    Args:
        file_content: Raw bytes of the Excel file
        
    Returns:
        List of dicts, each with keys: 'name', 'email', 'phone' (if present)
    """
    # openpyxl is imported here to avoid import errors if not installed
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel file support. Install with: pip install openpyxl")
    
    # Load workbook from bytes
    workbook = load_workbook(filename=io.BytesIO(file_content), read_only=True)
    sheet = workbook.active  # Get the first (active) sheet
    
    participants = []
    headers = []
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if row_idx == 0:
            # First row is headers - normalize them
            headers = [normalize_header(cell) for cell in row]
            continue
        
        # Build a dict from this row using headers
        row_dict = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                header = headers[col_idx]
                row_dict[header] = cell
        
        participant = {
            'name': str(row_dict.get('name', '')).strip() if row_dict.get('name') else None,
            'email': str(row_dict.get('email', '')).strip() if row_dict.get('email') else None,
            'phone': str(row_dict.get('phone', '')).strip() if row_dict.get('phone') else None,
        }
        participants.append(participant)
    
    workbook.close()
    return participants


def validate_participant_row(row_data: Dict[str, Any], row_number: int) -> Optional[str]:
    """
    Validate a single participant row.
    
    Args:
        row_data: Dict with 'name' and 'email' keys
        row_number: Row number for error reporting (1-based)
        
    Returns:
        Error message string if invalid, None if valid
    """
    name = row_data.get('name')
    email = row_data.get('email')
    
    # Both name and email must be present
    if not name and not email:
        return f"Row {row_number}: Empty name and email"
    if not name:
        return f"Row {row_number}: Missing name"
    if not email:
        return f"Row {row_number}: Missing email"
    
    return None  # Valid row

